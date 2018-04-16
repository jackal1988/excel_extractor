#!/usr/bin/env python3
# -*- coding: utf-8 -*-

'read and write NewEnergyCarCatalogExcel.py '

__author__ = 'RickyS'

import openpyxl, pprint, sys
from openpyxl.utils import get_column_letter
from tkinter.filedialog import askopenfilename

# from openpyxl import Workbook
# os.chdir(r"C:\Users\Administrator\Desktop")
# ---------------------Load workbook and worksheet----------------------------------------
print('正在打开工作簿......\n')
fname = askopenfilename()
wb = openpyxl.load_workbook(fname)
sheetAll = wb.sheetnames

BatchNum = int(input('工作簿 ' + str(fname) + ' 包含 ' + str(sheetAll) + ' 工作表\n\n请输入批次号\n'))

for eachSheetName in sheetAll:
# eachSheetName = 'Table 10'
    sheet = wb[eachSheetName]
    print('开始读取工作表 '+ '\"' + str(eachSheetName) + '\"')
    # ------------------------customize the rectangle area you want-----------------------

    startCell = input('请输入起始单元格名称（字母+数字）\n')
    endCell = input('请输入结束单元格名称（字母+数字）\n')
    recArea = sheet[startCell:endCell]  # recArea type is tuple
    # listRecArea = list(recArea)
    # print(recArea[0][1].value)
    # -----------------------Loop through recArea -----------------------------------------------

    lsEachDictKey = []
    lsEachDictValueAll = []
    for rowNum in range(0, len(recArea)):  # Iterate all row in recArea
        if rowNum == 0:  # save 1st line as key of dictionary
            for eachCell in recArea[rowNum]:
                if type(eachCell.value) == int:  # attribute column must be selected
                    print('属性列不在选择范围内，请重选。\n')
                    sys.exit()
                lsEachDictKey.append(eachCell.value)
            for m in range(0, len(lsEachDictKey)):  # standardize Key name, e.g. delete \n, replace () to （）, etc.
                if lsEachDictKey[m] is not None:
                    lsEachDictKey[m] = lsEachDictKey[m].replace(' ', '')
                    lsEachDictKey[m] = lsEachDictKey[m].replace('\n', '')
                    lsEachDictKey[m] = lsEachDictKey[m].replace('(', '（')
                    lsEachDictKey[m] = lsEachDictKey[m].replace(')', '）')
                    lsEachDictKey[m] = lsEachDictKey[m].replace('动力蓄电池总质量', '动力蓄电池组总质量')
                    lsEachDictKey[m] = lsEachDictKey[m].replace('动力蓄电池总能量', '动力蓄电池组总能量')
                    lsEachDictKey[m] = lsEachDictKey[m].replace('汽车企业名称', '汽车生产企业名称')
                else:
                    continue
            # tupEachDictKey = tuple(lsEachDictKey)
        else:
            for eachCell in recArea[rowNum]:
                lsEachDictValueAll.append(eachCell.value)
            for m in range(0, len(lsEachDictValueAll)):  # delete '\n' in lsEachDictValueAll.
                if lsEachDictValueAll[m] is not None and type(lsEachDictValueAll[m]) == str:  # 'replace()' method only
                    # effective towards str type data.
                    lsEachDictValueAll[m] = lsEachDictValueAll[m].replace('\n', '')
                    lsEachDictValueAll[m] = lsEachDictValueAll[m].replace(' ', '')
                else:
                    continue

        # dictData.setdefault(tupEachDictKey,lsEachDictValue)
    # pprint.pprint(lsEachDictKey)
    # Regroup lsEachDictValue-------------------------------------
    lsEachDictValueAllRegroup = []
    for c in range(0, len(lsEachDictKey)):
        n = c
        while n < len(lsEachDictValueAll):
            lsEachDictValueAllRegroup.append(lsEachDictValueAll[n])
            n += len(lsEachDictKey)
        c += 1
    # pprint.pprint(lsEachDictValueAllRegroup)
    stepLen = len(recArea) - 1
    # Auto fill up none company name value --------------------------------------
    for m in range(stepLen, 3 * stepLen):  # This method may not be flawless.
        if lsEachDictValueAllRegroup[m] is None:
            lsEachDictValueAllRegroup[m] = lsEachDictValueAllRegroup[m - 1]
    # pprint.pprint(lsEachDictValueAllRegroup)

    # Build dictionary data structure-----------------------------------------------------------
    dictData = {}
    for n in range(0, len(lsEachDictKey)):
        dictData.setdefault(lsEachDictKey[n], lsEachDictValueAllRegroup[n * stepLen:(n + 1) * stepLen])
    for eachKey in list(dictData.keys()):  # pop out 'None' key value
        # 被遍历的对象，在被遍历时其数据结构不能更改！！！所以用list(dictData.keys())代替dictData()
        if eachKey is None:
            dictData.pop(eachKey)
    for eachKey in list(dictData.keys()):  # ensure 'None' value not exist, in case not iterable.
        for eachValue in list(dictData[eachKey]):
            if eachValue is None:
                i = dictData[eachKey].index(eachValue)
                dictData[eachKey][i] = 'NoneValue 空值'
            else:
                continue
    # pprint.pprint(dictData)
    print('读取数据完毕，开始写入...\n')
    # -------------------------------write dictData to .xlsx file
    destFileName = 'catalog_of_new_energy_car.xlsx'
    wbUpdate = openpyxl.load_workbook(destFileName)
    wsUpdate = wbUpdate.active
    # ----------------------------------------------------------------------------
    rowNumMax = wsUpdate.max_row  # auto adding towards '序号'
    lsWorkBookBatch = []
    for n in range(0, len(dictData['序号'])):
        counter = 0
        if type(dictData['序号'][n]) != int:
            dictData['序号'][n] = 'NoneValue 空值'
            n += 1
            counter += 1
        else:
            dictData['序号'][n] = dictData['序号'][n] + counter + rowNumMax - 1

    lsWs1stRowValueAll = []  # iterate all contents in 1st. row of wsUpdate EVERY TIME, in case adding new item in the
    #  future
    for m in range(1, wsUpdate.max_column + 1):
        lsWs1stRowValueAll.append(wsUpdate[get_column_letter(m) + '1'].value)
    tupWs1stRowValueAll = tuple(lsWs1stRowValueAll)

    for eachKey in dictData.keys():  # write dictData to current worksheet
        if eachKey in tupWs1stRowValueAll:
            pinPoint = tupWs1stRowValueAll.index(eachKey)
            for n in range(0, len(dictData[eachKey])):
                wsUpdate.cell(row=rowNumMax + n + 1, column=pinPoint + 1).value = dictData[eachKey][n]
    #  --------------------auto fill '批次' '技术类型' '车辆用途类别'
    pinPoint2 = tupWs1stRowValueAll.index('批次')
    pinPoint3 = tupWs1stRowValueAll.index('技术类型')
    pinPoint4 = tupWs1stRowValueAll.index('车辆用途类别')
    for n in range(0, len(dictData['序号'])):  # auto fill up column: '批次' '技术类型' '车辆用途类别'
        wsUpdate.cell(row=rowNumMax + n + 1, column=pinPoint2 + 1).value = BatchNum
        if '纯电动续驶里程（km）' in dictData and '燃料消耗量（L/100km）' not in dictData and '燃料电池系统额定功率（kW）' not in dictData:
            wsUpdate.cell(row=rowNumMax + n + 1, column=pinPoint3 + 1).value = '纯电动'
        elif '纯电动续驶里程（km）' in dictData and '燃料消耗量（L/100km）' in dictData and '燃料电池系统额定功率（kW）' not in dictData:
            wsUpdate.cell(row=rowNumMax + n + 1, column=pinPoint3 + 1).value = '插电式混合动力'
        else:
            wsUpdate.cell(row=rowNumMax + n + 1, column=pinPoint3 + 1).value = '燃料电池'

        if '通用名称' in dictData:
            wsUpdate.cell(row=rowNumMax + n + 1, column=pinPoint4 + 1).value = '乘用车'
        elif '产品名称' in dictData and '客车' in dictData['产品名称'][n]:
            wsUpdate.cell(row=rowNumMax + n + 1, column=pinPoint4 + 1).value = '客车'
        else:
            wsUpdate.cell(row=rowNumMax + n + 1, column=pinPoint4 + 1).value = '客车/专用车/货车'

    wbUpdate.save(filename=destFileName)
    print('工作表' + str(eachSheetName) + ' 数据写入完毕。\n')
print('第 ' + str(BatchNum) + ' 批新能源车目录已录入完毕。\n')


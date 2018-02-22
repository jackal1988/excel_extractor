#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# readCensusExcel.py - calculate population by SLQ

import openpyxl, pprint
from tkinter.filedialog import askopenfilename
# from openpyxl import Workbook


# from openpyxl.cell.cell import column_index_from_string

# os.chdir(r"C:\Users\Administrator\Desktop")
print('Opening workbook...')
fname = askopenfilename()
wb = openpyxl.load_workbook(fname)
sheet = wb.get_sheet_by_name('Table 4')
# ------------create a new workbook to store the selected data
# wbTemp = Workbook()
# sheetTemp = wbTemp.create_sheet('newEnergyCarList', 0)
# print(type(sheet))
# ------------------------#customize the rectangle area you want-----------------------
# startRow    = int(input("enter the start row number you want:"))
# startColumn = input("enter the start column letter you want:")
# endRow      = int(input("enter the end row number you want:"))
# endColumn   = input("enter the end column letter you want:")
# startCell = startColumn + str(startRow)
# endCell   = endColumn + str(endRow)
startCell = 'A2'
endCell = 'J55'
recArea = sheet[startCell:endCell]  # recArea type is tuple
# listRecArea = list(recArea)
# print(recArea[0][1].value)
# Loop through recArea --------------------------------------------------------------------------------------

lsEachDictKey = []
lsEachDictValueAll = []
for rowNum in range(0,len(recArea)):                     # Iterate all row in recArea
    if rowNum == 0: # save 1st line as key
        for eachCell in recArea[rowNum]:
                lsEachDictKey.append(eachCell.value)
        for m in range(0,len(lsEachDictKey)):  # delete '\n' in lsEachDictKey
            if lsEachDictKey[m] != None:
                lsEachDictKey[m] = lsEachDictKey[m].replace('\n','')
            else:
                continue
        tupEachDictKey = tuple(lsEachDictKey)
    else:
        for eachCell in recArea[rowNum]:
            lsEachDictValueAll.append(eachCell.value)
        for m in range(0,len(lsEachDictValueAll)):  # delete '\n' in lsEachDictValueAll.
            if lsEachDictValueAll[m] != None and type(lsEachDictValueAll[m]) == str: # 'replace()' method only effective towards str type data.
                lsEachDictValueAll[m] = lsEachDictValueAll[m].replace('\n','')
            else:
                continue

    # dictData.setdefault(tupEachDictKey,lsEachDictValue)
# pprint.pprint(lsEachDictKey)
# Regroup lsEachDictValue-------------------------------------
lsEachDictValueAllRegroup = []
for c in range(0,len(lsEachDictKey)):
    n = c
    while n < len(lsEachDictValueAll):
        lsEachDictValueAllRegroup.append(lsEachDictValueAll[n])
        n += len(lsEachDictKey)
    c += 1
# pprint.pprint(lsEachDictValueAllRegroup)
stepLen = len(recArea)-1
# Auto fill up none company name value cell --------------------------------------
for m in range(stepLen,2*stepLen):   # This method may not be flawless.
    if lsEachDictValueAllRegroup[m] == None:
        lsEachDictValueAllRegroup[m] = lsEachDictValueAllRegroup[m-1]
# pprint.pprint(lsEachDictValueAllRegroup)

# Dictionary data structure-----------------------------------------------------------
dictData = {}
for n in range(0,len(lsEachDictKey)):
    dictData.setdefault(lsEachDictKey[n],lsEachDictValueAllRegroup[n*stepLen:(n+1)*stepLen])
dictData.pop(None)  # pop out None key-value
pprint.pprint(dictData)


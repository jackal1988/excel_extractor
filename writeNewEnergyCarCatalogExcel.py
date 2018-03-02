#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# writeNewEnergyCarCatalogExcel.py  by SLQ

import os
os.chdir('C:\\Users\\shaolqi\\PycharmProjects\\excel_extractor')

import openpyxl, newCatalog4NEC
from openpyxl.utils import get_column_letter
# ---------------------create a new workbook to store newCatalog4NEC.py
destFileName = 'catalog_of_new_energy_car.xlsx'
wbUpdate = openpyxl.load_workbook(destFileName)
wsUpdate = wbUpdate.active
# ----------------------------------------------------------------------------
rowNumMax = wsUpdate.max_row  # auto adding towards '序号'
for n in range(0, len(newCatalog4NEC.allData['序号'])):
    newCatalog4NEC.allData['序号'][n]= newCatalog4NEC.allData['序号'][n] + rowNumMax - 1


lsWs1stRowValueAll = []  # iterate all contents in 1st. row of wsUpdate EVERY TIME, in case adding new item in the
#  future
for m in range(1, wsUpdate.max_column + 1):
    lsWs1stRowValueAll.append(wsUpdate[get_column_letter(m) + '1'].value)
tupWs1stRowValueAll = tuple(lsWs1stRowValueAll)
print(tupWs1stRowValueAll)

for eachKey in newCatalog4NEC.allData.keys():  # write dictData to current worksheet
    if eachKey in tupWs1stRowValueAll:
        pinPoint = tupWs1stRowValueAll.index(eachKey)
        for n in range(0, len(newCatalog4NEC.allData[eachKey])):
            wsUpdate.cell(row =rowNumMax + n + 1, column =pinPoint + 1).value = newCatalog4NEC.allData[eachKey][n]

wbUpdate.save(filename=destFileName)
